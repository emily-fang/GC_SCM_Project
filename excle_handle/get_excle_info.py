from  openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color
def get_excel_request_data(file_path,row,col,maxcol):
      """

      :param path: 文件相对路径，row首行参数名称和第几行（row）的参数值拼接,col是从第几列开始参数拼接,maxcol到第几列结束
      :return: 返回文件简称，请求参数
      """
      open_excel = load_workbook(file_path)
      excel = open_excel.active

      request_data = dict()
      short_name = excel.cell(row, 1).value

      for i in range(col,maxcol):
            request_data[excel.cell(1,i).value] = excel.cell(row,i).value

      return  short_name,request_data

def  write_short_name(path,short_names,col):
      """
      :param path: 要写入信息的路径，文件名
      :param short_names: 要写入的文件名称（缩写）
      :param col: 文件前缀，写到第几列
      :return:
      """
      open_excel = load_workbook(path)
      excel = open_excel.active
      for i in range(2,len(short_names)+2):
            excel.cell(i,col).value = short_names[i-2]
      open_excel.save(path)

def  write_reslut(path,reslut,row,col):
      """
      往指定的单元格，写入指定的结果
     """
      open_excel = load_workbook(path)
      excel = open_excel.active

      excel.cell(row,col).value=reslut
      open_excel.save(path)

def write_result_to_xlsx(path, row, col, result):
    """
    将结果写入xlsx文件
    """
    xlsx_file = load_workbook(path)
    sheet_one = xlsx_file.active
    if result.upper() =="PASS":
          result_color = "99FF66"
    else:
          result_color = "ffd3d3"

    font_color = PatternFill("solid", fgColor=result_color)
    sheet_one.cell(row, col).fill = font_color
    xlsx_file.save(path)

if __name__ == '__main__':
    test_num,reqest_data =get_excel_request_data("../scm_data/scm_request_data.xlsx", 2, 2, 7)
    print(reqest_data)
    # get_excel_request_data("../file/data.xlsx",2,3,7)
